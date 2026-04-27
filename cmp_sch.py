"""
schema_comparator.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Hive DDL  ↔  Snowflake INFORMATION_SCHEMA schema comparator.

Usage
-----
  python schema_comparator.py --ddl-dir ./ddl_files \
      --account myaccount --user myuser --password mypass \
      --warehouse COMPUTE_WH --database MY_DB --schema MY_SCHEMA \
      [--parallel] [--workers 8] [--output report.xlsx]

Flags
-----
  --parallel          Run DDL parsing in parallel (default: serial)
  --workers N         Thread-pool size when --parallel is set (default: cpu_count)
  --output FILE       Excel output path (default: schema_diff_report.xlsx)
  --log-level LEVEL   DEBUG | INFO | WARNING (default: INFO)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import argparse
import logging
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import snowflake.connector
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("schema_comparator")


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ColumnInfo:
    name: str
    data_type: str          # normalised base type
    raw_type: str           # original string from DDL / SF
    is_complex: bool = False
    char_length: Optional[int] = None
    numeric_precision: Optional[int] = None
    numeric_scale: Optional[int] = None
    ordinal_position: Optional[int] = None
    is_partitioned: bool = False    # Hive-only flag
    nullable: Optional[bool] = None


@dataclass
class TableSchema:
    table_name: str
    columns: Dict[str, ColumnInfo] = field(default_factory=dict)
    source_file: Optional[str] = None          # DDL path (Hive only)
    parse_error: Optional[str] = None


@dataclass
class ColumnDiff:
    table: str
    column: str
    hive_type: str
    sf_type: str
    issue: str          # see ISSUE_* constants below
    hive_raw: str = ""
    sf_raw: str = ""
    is_complex: bool = False
    source_file: str = ""


# Issue taxonomy
ISSUE_TYPE_MISMATCH       = "TYPE_MISMATCH"
ISSUE_MISSING_IN_SF       = "MISSING_IN_SNOWFLAKE"
ISSUE_MISSING_IN_HIVE     = "MISSING_IN_HIVE"
ISSUE_TABLE_NOT_IN_SF     = "TABLE_NOT_IN_SNOWFLAKE"
ISSUE_TABLE_NOT_IN_HIVE   = "TABLE_NOT_IN_HIVE"
ISSUE_COMPLEX_REVIEW      = "COMPLEX_TYPE_NEEDS_REVIEW"
ISSUE_PARSE_ERROR         = "DDL_PARSE_ERROR"
ISSUE_OK                  = "MATCH"


# ─────────────────────────────────────────────────────────────────────────────
# Type normalisation
# ─────────────────────────────────────────────────────────────────────────────

# Hive base-type → canonical Snowflake equivalent
HIVE_TO_SF_TYPE: Dict[str, str] = {
    # Strings
    "STRING":       "TEXT",
    "VARCHAR":      "TEXT",
    "CHAR":         "TEXT",
    # Integers
    "TINYINT":      "NUMBER",
    "SMALLINT":     "NUMBER",
    "INT":          "NUMBER",
    "INTEGER":      "NUMBER",
    "BIGINT":       "NUMBER",
    # Floating point
    "FLOAT":        "FLOAT",
    "DOUBLE":       "FLOAT",
    "DOUBLE PRECISION": "FLOAT",
    # Fixed-precision
    "DECIMAL":      "NUMBER",
    "NUMERIC":      "NUMBER",
    # Boolean
    "BOOLEAN":      "BOOLEAN",
    # Date / time
    "DATE":         "DATE",
    "TIMESTAMP":    "TIMESTAMP_NTZ",
    "INTERVAL":     "TEXT",        # no native SF equivalent
    # Binary
    "BINARY":       "BINARY",
    "VARBINARY":    "BINARY",
    # Complex — all map to VARIANT in Snowflake
    "ARRAY":        "VARIANT",
    "MAP":          "VARIANT",
    "STRUCT":       "VARIANT",
    "UNIONTYPE":    "VARIANT",
}

# Snowflake synonyms → canonical name (SF returns these from INFORMATION_SCHEMA)
SF_TYPE_ALIASES: Dict[str, str] = {
    "TEXT":             "TEXT",
    "VARCHAR":          "TEXT",
    "CHARACTER VARYING":"TEXT",
    "STRING":           "TEXT",
    "CHAR":             "TEXT",
    "CHARACTER":        "TEXT",
    "NUMBER":           "NUMBER",
    "DECIMAL":          "NUMBER",
    "NUMERIC":          "NUMBER",
    "INT":              "NUMBER",
    "INTEGER":          "NUMBER",
    "BIGINT":           "NUMBER",
    "SMALLINT":         "NUMBER",
    "TINYINT":          "NUMBER",
    "BYTEINT":          "NUMBER",
    "FLOAT":            "FLOAT",
    "FLOAT4":           "FLOAT",
    "FLOAT8":           "FLOAT",
    "DOUBLE":           "FLOAT",
    "DOUBLE PRECISION": "FLOAT",
    "REAL":             "FLOAT",
    "BOOLEAN":          "BOOLEAN",
    "DATE":             "DATE",
    "DATETIME":         "TIMESTAMP_NTZ",
    "TIME":             "TIME",
    "TIMESTAMP":        "TIMESTAMP_NTZ",
    "TIMESTAMP_NTZ":    "TIMESTAMP_NTZ",
    "TIMESTAMP_LTZ":    "TIMESTAMP_LTZ",
    "TIMESTAMP_TZ":     "TIMESTAMP_TZ",
    "VARIANT":          "VARIANT",
    "OBJECT":           "VARIANT",
    "ARRAY":            "VARIANT",
    "BINARY":           "BINARY",
    "VARBINARY":        "BINARY",
}

COMPLEX_HIVE_TYPES = {"ARRAY", "MAP", "STRUCT", "UNIONTYPE"}


def _extract_base_type(type_str: str) -> str:
    """'DECIMAL(10,2)' → 'DECIMAL',  'ARRAY<STRING>' → 'ARRAY'"""
    return re.split(r"[<(\s]", type_str.strip())[0].upper()


def normalise_hive_type(raw: str) -> Tuple[str, bool]:
    """Return (canonical_type, is_complex)."""
    base = _extract_base_type(raw)
    is_complex = base in COMPLEX_HIVE_TYPES
    canonical = HIVE_TO_SF_TYPE.get(base, base)
    return canonical, is_complex


def normalise_sf_type(raw: str) -> str:
    """Return canonical Snowflake type."""
    base = _extract_base_type(raw)
    return SF_TYPE_ALIASES.get(base, base)


# ─────────────────────────────────────────────────────────────────────────────
# Hive DDL parser
# ─────────────────────────────────────────────────────────────────────────────

# Matches:  col_name  TYPE  [COMMENT '...']  [,]
_COL_RE = re.compile(
    r"^\s*`?(\w+)`?\s+"                      # column name
    r"((?:ARRAY|MAP|STRUCT|UNIONTYPE)"        # complex …
    r"(?:<[^)]+>)?"                           # … with angle-bracket args
    r"|[A-Z][A-Z0-9 ]*"                       # simple type (may have space, e.g. DOUBLE PRECISION)
    r"(?:\(\s*\d+(?:\s*,\s*\d+)?\s*\))?)"    # optional (p) or (p,s)
    r".*",
    re.IGNORECASE,
)


def _strip_comments(text: str) -> str:
    """Remove SQL line comments (--) and block comments (/* */)."""
    text = re.sub(r"/\*.*?\*/", " ", text, flags=re.DOTALL)
    text = re.sub(r"--[^\n]*", "", text)
    return text


def parse_hive_ddl(sql_file: Path) -> TableSchema:
    try:
        raw = sql_file.read_text(encoding="utf-8", errors="replace")
        text = _strip_comments(raw).upper()

        # ── Table name ──────────────────────────────────────────────────────
        tbl_match = re.search(
            r"CREATE\s+(?:EXTERNAL\s+)?TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:\w+\.)?(\w+)",
            text,
        )
        if not tbl_match:
            return TableSchema(
                table_name=sql_file.stem.upper(),
                source_file=str(sql_file),
                parse_error="Could not find CREATE TABLE statement",
            )
        table_name = tbl_match.group(1)

        # ── Column block ─────────────────────────────────────────────────────
        # Find the opening paren of the column list
        paren_start = text.find("(", tbl_match.end())
        if paren_start == -1:
            return TableSchema(table_name=table_name, source_file=str(sql_file),
                               parse_error="No column block found")

        # Walk to the matching closing paren (depth-aware)
        depth = 0
        paren_end = paren_start
        for i, ch in enumerate(text[paren_start:], start=paren_start):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    paren_end = i
                    break

        col_block = text[paren_start + 1: paren_end]

        # ── Partitioned columns ──────────────────────────────────────────────
        part_cols: set = set()
        part_match = re.search(
            r"PARTITIONED\s+BY\s*\((.+?)\)", text[paren_end:], re.DOTALL
        )
        if part_match:
            for line in part_match.group(1).splitlines():
                m = _COL_RE.match(line.strip().rstrip(","))
                if m:
                    part_cols.add(m.group(1).upper())

        # ── Parse columns ────────────────────────────────────────────────────
        columns: Dict[str, ColumnInfo] = {}
        ordinal = 0
        for line in col_block.splitlines():
            line = line.strip().rstrip(",")
            if not line or line.startswith("PRIMARY") or line.startswith("CONSTRAINT"):
                continue
            m = _COL_RE.match(line)
            if not m:
                continue
            col_name = m.group(1).upper()
            raw_type = m.group(2).strip()
            canonical, is_complex = normalise_hive_type(raw_type)

            # Precision / scale extraction for DECIMAL(p,s)
            prec, scale = None, None
            ps_match = re.search(r"\((\d+)(?:,\s*(\d+))?\)", raw_type)
            if ps_match:
                prec = int(ps_match.group(1))
                scale = int(ps_match.group(2)) if ps_match.group(2) else None

            ordinal += 1
            columns[col_name] = ColumnInfo(
                name=col_name,
                data_type=canonical,
                raw_type=raw_type,
                is_complex=is_complex,
                numeric_precision=prec,
                numeric_scale=scale,
                ordinal_position=ordinal,
                is_partitioned=(col_name in part_cols),
            )

        if not columns:
            return TableSchema(table_name=table_name, source_file=str(sql_file),
                               parse_error="No columns parsed from DDL")

        log.debug("Parsed %s → %d columns", table_name, len(columns))
        return TableSchema(table_name=table_name, columns=columns,
                           source_file=str(sql_file))

    except Exception as exc:
        log.warning("Error parsing %s: %s", sql_file, exc)
        return TableSchema(
            table_name=sql_file.stem.upper(),
            source_file=str(sql_file),
            parse_error=str(exc),
        )


# ─────────────────────────────────────────────────────────────────────────────
# Snowflake schema extraction
# ─────────────────────────────────────────────────────────────────────────────

def fetch_snowflake_schemas(
    conn: snowflake.connector.SnowflakeConnection,
    database: str,
    schema: str,
) -> Dict[str, TableSchema]:
    """
    Pull ALL columns for the target schema in a single query.
    Returns {TABLE_NAME: TableSchema}.
    """
    log.info("Fetching Snowflake schemas for %s.%s …", database, schema)
    query = f"""
        SELECT
            TABLE_NAME,
            COLUMN_NAME,
            DATA_TYPE,
            CHARACTER_MAXIMUM_LENGTH,
            NUMERIC_PRECISION,
            NUMERIC_SCALE,
            IS_NULLABLE,
            ORDINAL_POSITION
        FROM {database}.INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = '{schema.upper()}'
        ORDER BY TABLE_NAME, ORDINAL_POSITION
    """
    cur = conn.cursor()
    cur.execute(query)
    rows = cur.fetchall()

    sf_tables: Dict[str, TableSchema] = {}
    for (tbl, col, dtype, char_len, num_prec, num_scale, nullable, ordinal) in rows:
        if tbl not in sf_tables:
            sf_tables[tbl] = TableSchema(table_name=tbl)
        canonical = normalise_sf_type(dtype)
        sf_tables[tbl].columns[col] = ColumnInfo(
            name=col,
            data_type=canonical,
            raw_type=dtype,
            is_complex=(canonical == "VARIANT"),
            char_length=char_len,
            numeric_precision=num_prec,
            numeric_scale=num_scale,
            nullable=(nullable == "YES"),
            ordinal_position=ordinal,
        )

    log.info("Fetched %d tables from Snowflake.", len(sf_tables))
    return sf_tables


# ─────────────────────────────────────────────────────────────────────────────
# Schema comparison engine
# ─────────────────────────────────────────────────────────────────────────────

def compare_table(
    hive: TableSchema,
    sf_tables: Dict[str, TableSchema],
) -> List[ColumnDiff]:
    diffs: List[ColumnDiff] = []

    if hive.parse_error:
        diffs.append(ColumnDiff(
            table=hive.table_name, column="—",
            hive_type=f"PARSE ERROR: {hive.parse_error}",
            sf_type="—", issue=ISSUE_PARSE_ERROR,
            source_file=hive.source_file or "",
        ))
        return diffs

    if hive.table_name not in sf_tables:
        diffs.append(ColumnDiff(
            table=hive.table_name, column="—",
            hive_type="(all columns)", sf_type="—",
            issue=ISSUE_TABLE_NOT_IN_SF,
            source_file=hive.source_file or "",
        ))
        return diffs

    sf = sf_tables[hive.table_name]
    all_cols = sorted(set(hive.columns) | set(sf.columns))

    for col in all_cols:
        h_info = hive.columns.get(col)
        s_info = sf.columns.get(col)

        if h_info is None:
            diffs.append(ColumnDiff(
                table=hive.table_name, column=col,
                hive_type="—", sf_type=s_info.data_type,
                hive_raw="—", sf_raw=s_info.raw_type,
                issue=ISSUE_MISSING_IN_HIVE,
                source_file=hive.source_file or "",
            ))
        elif s_info is None:
            issue = ISSUE_COMPLEX_REVIEW if h_info.is_complex else ISSUE_MISSING_IN_SF
            diffs.append(ColumnDiff(
                table=hive.table_name, column=col,
                hive_type=h_info.data_type, sf_type="—",
                hive_raw=h_info.raw_type, sf_raw="—",
                issue=issue, is_complex=h_info.is_complex,
                source_file=hive.source_file or "",
            ))
        elif h_info.data_type != s_info.data_type:
            # Complex types mapped to VARIANT are expected — flag for review not error
            issue = (ISSUE_COMPLEX_REVIEW if h_info.is_complex or s_info.is_complex
                     else ISSUE_TYPE_MISMATCH)
            diffs.append(ColumnDiff(
                table=hive.table_name, column=col,
                hive_type=h_info.data_type, sf_type=s_info.data_type,
                hive_raw=h_info.raw_type, sf_raw=s_info.raw_type,
                issue=issue, is_complex=(h_info.is_complex or s_info.is_complex),
                source_file=hive.source_file or "",
            ))
        # else: MATCH — only log at DEBUG level
        else:
            log.debug("✓ %s.%s  [%s]", hive.table_name, col, h_info.data_type)

    return diffs


# ─────────────────────────────────────────────────────────────────────────────
# Parallel / serial orchestration
# ─────────────────────────────────────────────────────────────────────────────

def run_comparison(
    ddl_dir: str,
    sf_tables: Dict[str, TableSchema],
    parallel: bool = False,
    workers: int = None,
) -> Tuple[List[ColumnDiff], List[TableSchema], List[str]]:
    """
    Parse all DDL files and compare against Snowflake schemas.

    Returns
    -------
    diffs         : all ColumnDiff records
    hive_schemas  : all parsed TableSchema objects
    sf_only_tables: tables in SF but not found in any DDL file
    """
    ddl_files = list(Path(ddl_dir).rglob("*.sql"))
    if not ddl_files:
        log.error("No .sql files found in %s", ddl_dir)
        sys.exit(1)

    log.info("Found %d DDL files.  Mode: %s", len(ddl_files),
             f"PARALLEL (workers={workers or 'auto'})" if parallel else "SERIAL")

    hive_schemas: List[TableSchema] = []

    # ── Parse DDL files ──────────────────────────────────────────────────────
    t0 = time.perf_counter()
    if parallel:
        max_workers = workers or os.cpu_count()
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {pool.submit(parse_hive_ddl, f): f for f in ddl_files}
            for future in as_completed(futures):
                hive_schemas.append(future.result())
    else:
        for f in ddl_files:
            hive_schemas.append(parse_hive_ddl(f))

    parse_time = time.perf_counter() - t0
    log.info("Parsed %d DDL files in %.2fs.", len(hive_schemas), parse_time)

    # ── Compare ──────────────────────────────────────────────────────────────
    all_diffs: List[ColumnDiff] = []
    hive_table_names = {s.table_name for s in hive_schemas}

    for hs in hive_schemas:
        all_diffs.extend(compare_table(hs, sf_tables))

    # Tables in Snowflake with no corresponding DDL file
    sf_only = [t for t in sf_tables if t not in hive_table_names]
    for tbl in sf_only:
        all_diffs.append(ColumnDiff(
            table=tbl, column="—",
            hive_type="—", sf_type="(all columns)",
            issue=ISSUE_TABLE_NOT_IN_HIVE,
        ))

    return all_diffs, hive_schemas, sf_only


# ─────────────────────────────────────────────────────────────────────────────
# Excel report builder
# ─────────────────────────────────────────────────────────────────────────────

# Colour palette
C_RED    = "FFDDDD"
C_ORANGE = "FFE8CC"
C_YELLOW = "FFFACC"
C_GREEN  = "DDFFD8"
C_BLUE   = "D8EEFF"
C_PURPLE = "EFD8FF"
C_GREY   = "F2F2F2"
C_WHITE  = "FFFFFF"
C_HEADER = "2D5FA6"
C_HEADER_FG = "FFFFFF"

ISSUE_STYLE = {
    ISSUE_TYPE_MISMATCH:     (C_RED,    "❌"),
    ISSUE_MISSING_IN_SF:     (C_ORANGE, "⚠️"),
    ISSUE_MISSING_IN_HIVE:   (C_YELLOW, "⚠️"),
    ISSUE_TABLE_NOT_IN_SF:   (C_RED,    "🔴"),
    ISSUE_TABLE_NOT_IN_HIVE: (C_BLUE,   "🔵"),
    ISSUE_COMPLEX_REVIEW:    (C_PURPLE, "🔷"),
    ISSUE_PARSE_ERROR:       (C_GREY,   "⚙️"),
    ISSUE_OK:                (C_GREEN,  "✅"),
}

def _hdr_fill(colour: str = C_HEADER) -> PatternFill:
    return PatternFill("solid", fgColor=colour)

def _cell_fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)

def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_header_row(ws, headers: List[str], row: int = 1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=C_HEADER_FG, size=11)
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 22

def _auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def build_excel_report(
    diffs: List[ColumnDiff],
    hive_schemas: List[TableSchema],
    sf_tables: Dict[str, TableSchema],
    output_path: str,
):
    log.info("Building Excel report → %s", output_path)
    wb = Workbook()

    # ── Categorise diffs ─────────────────────────────────────────────────────
    by_issue: Dict[str, List[ColumnDiff]] = {k: [] for k in ISSUE_STYLE}
    for d in diffs:
        by_issue.setdefault(d.issue, []).append(d)

    # Matched tables (no diffs)
    hive_names = {s.table_name for s in hive_schemas}
    matched_tables = [t for t in hive_names
                      if t in sf_tables
                      and not any(d.table == t for d in diffs)]

    # ── Sheet 1: Executive Summary ───────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📊 Summary"

    total_tables = len(hive_schemas)
    total_sf     = len(sf_tables)
    match_count  = len(matched_tables)

    summary_data = [
        ("Metric", "Count"),
        ("DDL files parsed", total_tables),
        ("Snowflake tables found", total_sf),
        ("Tables fully matched ✅", match_count),
        ("", ""),
        ("Issue Type", "Count"),
        (f"❌ {ISSUE_TYPE_MISMATCH}", len(by_issue[ISSUE_TYPE_MISMATCH])),
        (f"⚠️  {ISSUE_MISSING_IN_SF}", len(by_issue[ISSUE_MISSING_IN_SF])),
        (f"⚠️  {ISSUE_MISSING_IN_HIVE}", len(by_issue[ISSUE_MISSING_IN_HIVE])),
        (f"🔴 {ISSUE_TABLE_NOT_IN_SF}", len(by_issue[ISSUE_TABLE_NOT_IN_SF])),
        (f"🔵 {ISSUE_TABLE_NOT_IN_HIVE}", len(by_issue[ISSUE_TABLE_NOT_IN_HIVE])),
        (f"🔷 {ISSUE_COMPLEX_REVIEW}", len(by_issue[ISSUE_COMPLEX_REVIEW])),
        (f"⚙️  {ISSUE_PARSE_ERROR}", len(by_issue[ISSUE_PARSE_ERROR])),
        ("", ""),
        ("Total mismatches", len(diffs)),
    ]

    for r, (label, value) in enumerate(summary_data, 1):
        c1 = ws_sum.cell(row=r, column=1, value=label)
        c2 = ws_sum.cell(row=r, column=2, value=value)
        if label in ("Metric", "Issue Type"):
            for c in (c1, c2):
                c.font = Font(bold=True, color=C_HEADER_FG)
                c.fill = _hdr_fill()
        elif label == "":
            pass
        else:
            c1.font = Font(bold=False)
            c2.alignment = Alignment(horizontal="center")
            # Colour rows by issue
            for issue, (colour, _) in ISSUE_STYLE.items():
                if issue in label:
                    c1.fill = _cell_fill(colour)
                    c2.fill = _cell_fill(colour)
                    break
        for c in (c1, c2):
            c.border = _thin_border()

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 14

    # ── Sheet 2: All Diffs ───────────────────────────────────────────────────
    ws_all = wb.create_sheet("🔍 All Issues")
    hdrs = ["Table", "Column", "Hive Raw Type", "Hive Canonical",
            "SF Raw Type", "SF Canonical", "Issue", "Is Complex", "DDL File"]
    _write_header_row(ws_all, hdrs)

    for r, d in enumerate(sorted(diffs, key=lambda x: (x.issue, x.table, x.column)), 2):
        colour, emoji = ISSUE_STYLE.get(d.issue, (C_WHITE, ""))
        row_vals = [
            d.table, d.column,
            d.hive_raw, d.hive_type,
            d.sf_raw, d.sf_type,
            f"{emoji} {d.issue}",
            "YES" if d.is_complex else "",
            Path(d.source_file).name if d.source_file else "",
        ]
        fill = _cell_fill(colour)
        for c, val in enumerate(row_vals, 1):
            cell = ws_all.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
    _auto_width(ws_all)
    _freeze(ws_all)

    # ── Sheets 3–8: One sheet per issue type ─────────────────────────────────
    issue_sheet_map = {
        ISSUE_TYPE_MISMATCH:     "❌ Type Mismatches",
        ISSUE_MISSING_IN_SF:     "⚠️ Missing in SF",
        ISSUE_MISSING_IN_HIVE:   "⚠️ Missing in Hive",
        ISSUE_TABLE_NOT_IN_SF:   "🔴 Tables Not in SF",
        ISSUE_TABLE_NOT_IN_HIVE: "🔵 Tables Not in Hive",
        ISSUE_COMPLEX_REVIEW:    "🔷 Complex Types",
        ISSUE_PARSE_ERROR:       "⚙️ Parse Errors",
    }

    for issue, sheet_name in issue_sheet_map.items():
        records = by_issue.get(issue, [])
        if not records:
            continue
        ws = wb.create_sheet(sheet_name)
        _write_header_row(ws, hdrs)
        colour, _ = ISSUE_STYLE[issue]
        fill = _cell_fill(colour)
        for r, d in enumerate(records, 2):
            for c, val in enumerate([
                d.table, d.column, d.hive_raw, d.hive_type,
                d.sf_raw, d.sf_type, d.issue,
                "YES" if d.is_complex else "",
                Path(d.source_file).name if d.source_file else "",
            ], 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = fill
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center")
        ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
        _auto_width(ws)
        _freeze(ws)

    # ── Sheet 9: Matched Tables ──────────────────────────────────────────────
    ws_ok = wb.create_sheet("✅ Matched Tables")
    _write_header_row(ws_ok, ["Table Name", "Column Count (Hive)", "Column Count (SF)"])
    green = _cell_fill(C_GREEN)
    for r, tname in enumerate(sorted(matched_tables), 2):
        for c, val in enumerate([
            tname,
            len(sf_tables.get(tname, TableSchema("")).columns),
            len(sf_tables.get(tname, TableSchema("")).columns),
        ], 1):
            cell = ws_ok.cell(row=r, column=c, value=val)
            cell.fill = green
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")
    _auto_width(ws_ok)
    _freeze(ws_ok, "A2")

    wb.save(output_path)
    log.info("Report saved: %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Compare Hive DDL schemas against Snowflake INFORMATION_SCHEMA.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--ddl-dir",    required=True,  help="Directory containing .sql DDL files")
    p.add_argument("--account",    required=True,  help="Snowflake account identifier")
    p.add_argument("--user",       required=True,  help="Snowflake username")
    p.add_argument("--password",   required=True,  help="Snowflake password")
    p.add_argument("--warehouse",  required=True,  help="Snowflake warehouse")
    p.add_argument("--database",   required=True,  help="Snowflake database")
    p.add_argument("--schema",     required=True,  help="Snowflake schema")
    p.add_argument("--role",       default=None,   help="Snowflake role (optional)")
    p.add_argument("--output",     default="schema_diff_report.xlsx",
                   help="Output Excel file path")
    p.add_argument("--parallel",   action="store_true",
                   help="Parse DDL files in parallel (faster for 1000+ files)")
    p.add_argument("--workers",    type=int, default=None,
                   help="Thread pool size for --parallel mode (default: cpu count)")
    p.add_argument("--log-level",  default="INFO",
                   choices=["DEBUG", "INFO", "WARNING"],
                   help="Logging verbosity")
    return p


def main():
    args = build_arg_parser().parse_args()
    log.setLevel(args.log_level)

    # ── Connect to Snowflake ─────────────────────────────────────────────────
    log.info("Connecting to Snowflake account: %s …", args.account)
    conn_params = dict(
        account=args.account,
        user=args.user,
        password=args.password,
        warehouse=args.warehouse,
        database=args.database,
        schema=args.schema,
    )
    if args.role:
        conn_params["role"] = args.role

    try:
        conn = snowflake.connector.connect(**conn_params)
    except Exception as exc:
        log.error("Snowflake connection failed: %s", exc)
        sys.exit(1)

    # ── Fetch Snowflake schemas ──────────────────────────────────────────────
    sf_tables = fetch_snowflake_schemas(conn, args.database, args.schema)
    conn.close()

    # ── Parse DDLs & compare ─────────────────────────────────────────────────
    t_start = time.perf_counter()
    diffs, hive_schemas, sf_only = run_comparison(
        ddl_dir=args.ddl_dir,
        sf_tables=sf_tables,
        parallel=args.parallel,
        workers=args.workers,
    )
    elapsed = time.perf_counter() - t_start

    # ── Print summary to console ─────────────────────────────────────────────
    print("\n" + "━" * 60)
    print(f"  SCHEMA COMPARISON COMPLETE  ({elapsed:.1f}s)")
    print("━" * 60)
    from collections import Counter
    counts = Counter(d.issue for d in diffs)
    for issue, (_, emoji) in ISSUE_STYLE.items():
        if counts[issue]:
            print(f"  {emoji}  {issue:<35} {counts[issue]:>5}")
    print("━" * 60)
    print(f"  Total mismatches: {len(diffs)}")
    print(f"  Matched tables:   "
          f"{len(hive_schemas) - len(set(d.table for d in diffs if d.issue not in (ISSUE_TABLE_NOT_IN_HIVE,)))}")
    print("━" * 60 + "\n")

    # ── Build Excel report ───────────────────────────────────────────────────
    build_excel_report(diffs, hive_schemas, sf_tables, args.output)
    print(f"  📄 Report → {args.output}\n")


if __name__ == "__main__":
    main()
